import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import subprocess
import platform

class ReportManager:
    def __init__(self, database):
        self.db = database
        
    def create_interface(self, parent):
        # Ana frame
        main_frame = ttk.Frame(parent)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Üst panel - Rapor seçenekleri
        top_frame = ttk.LabelFrame(main_frame, text="Rapor Seçenekleri", padding=10)
        top_frame.pack(fill='x', pady=(0, 10))
        
        # Rapor türü seçimi
        report_frame = ttk.Frame(top_frame)
        report_frame.pack(fill='x')
        
        ttk.Label(report_frame, text="Rapor Türü:").pack(side='left', padx=(0, 10))
        
        self.report_type = tk.StringVar(value="stock_summary")
        
        ttk.Radiobutton(report_frame, text="Stok Özeti", 
                       variable=self.report_type, value="stock_summary").pack(side='left', padx=5)
        ttk.Radiobutton(report_frame, text="Stok Hareketleri", 
                       variable=self.report_type, value="stock_movements").pack(side='left', padx=5)
        ttk.Radiobutton(report_frame, text="Düşük Stok", 
                       variable=self.report_type, value="low_stock").pack(side='left', padx=5)
        ttk.Radiobutton(report_frame, text="Değer Analizi", 
                       variable=self.report_type, value="value_analysis").pack(side='left', padx=5)
        
        # Tarih aralığı seçimi
        date_frame = ttk.Frame(top_frame)
        date_frame.pack(fill='x', pady=10)
        
        ttk.Label(date_frame, text="Tarih Aralığı:").pack(side='left', padx=(0, 10))
        
        self.date_range = tk.StringVar(value="all")
        
        ttk.Radiobutton(date_frame, text="Tümü", 
                       variable=self.date_range, value="all").pack(side='left', padx=5)
        ttk.Radiobutton(date_frame, text="Bugün", 
                       variable=self.date_range, value="today").pack(side='left', padx=5)
        ttk.Radiobutton(date_frame, text="Bu Hafta", 
                       variable=self.date_range, value="week").pack(side='left', padx=5)
        ttk.Radiobutton(date_frame, text="Bu Ay", 
                       variable=self.date_range, value="month").pack(side='left', padx=5)
        
        # Butonlar
        button_frame = ttk.Frame(top_frame)
        button_frame.pack(pady=10)
        
        ttk.Button(button_frame, text="Rapor Oluştur", 
                  command=self.generate_report).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Excel'e Aktar", 
                  command=self.export_to_excel).pack(side='left', padx=5)
        ttk.Button(button_frame, text="Yazdır", 
                  command=self.print_report).pack(side='left', padx=5)
        
        # Alt panel - Rapor görüntüleme alanı
        bottom_frame = ttk.LabelFrame(main_frame, text="Rapor Sonuçları", padding=10)
        bottom_frame.pack(fill='both', expand=True)
        
        # Notebook için farklı rapor türleri
        self.report_notebook = ttk.Notebook(bottom_frame)
        self.report_notebook.pack(fill='both', expand=True)
        
        # Tablo görünümü
        self.create_table_view()
        
        # Özet görünümü
        self.create_summary_view()
        
        # Başlangıçta rapor oluştur
        self.generate_report()
    
    def create_table_view(self):
        """Tablo görünümü oluştur"""
        table_frame = ttk.Frame(self.report_notebook)
        self.report_notebook.add(table_frame, text="Detaylı Tablo")
        
        # Tablo
        self.report_tree = ttk.Treeview(table_frame, show='headings', height=20)
        
        # Scrollbar
        table_scrollbar = ttk.Scrollbar(table_frame, orient='vertical', 
                                       command=self.report_tree.yview)
        self.report_tree.configure(yscrollcommand=table_scrollbar.set)
        
        self.report_tree.pack(side='left', fill='both', expand=True)
        table_scrollbar.pack(side='right', fill='y')
    
    def create_summary_view(self):
        """Özet görünümü oluştur"""
        summary_frame = ttk.Frame(self.report_notebook)
        self.report_notebook.add(summary_frame, text="Özet Bilgiler")
        
        # Özet kartları için ana frame
        cards_frame = ttk.Frame(summary_frame)
        cards_frame.pack(fill='x', pady=10)
        
        # Kart 1 - Toplam Ürün Sayısı
        self.create_summary_card(cards_frame, "Toplam Ürün", "0", "#3498db", 0, 0)
        
        # Kart 2 - Toplam Stok Değeri
        self.create_summary_card(cards_frame, "Toplam Değer", "0 ₺", "#2ecc71", 0, 1)
        
        # Kart 3 - Düşük Stoklu Ürün
        self.create_summary_card(cards_frame, "Düşük Stok", "0", "#e74c3c", 0, 2)
        
        # Kart 4 - Bu Ay Hareket
        self.create_summary_card(cards_frame, "Aylık Hareket", "0", "#f39c12", 1, 0)
        
        # Kart 5 - En Çok Satan
        self.create_summary_card(cards_frame, "En Çok Satan", "-", "#9b59b6", 1, 1)
        
        # Kart 6 - Kritik Seviye
        self.create_summary_card(cards_frame, "Kritik Seviye", "0", "#e67e22", 1, 2)
        
        # Detaylı bilgi alanı
        detail_frame = ttk.LabelFrame(summary_frame, text="Detaylı Analiz", padding=10)
        detail_frame.pack(fill='both', expand=True, pady=10)
        
        self.detail_text = tk.Text(detail_frame, height=15, wrap=tk.WORD)
        detail_scrollbar = ttk.Scrollbar(detail_frame, orient='vertical', 
                                        command=self.detail_text.yview)
        self.detail_text.configure(yscrollcommand=detail_scrollbar.set)
        
        self.detail_text.pack(side='left', fill='both', expand=True)
        detail_scrollbar.pack(side='right', fill='y')
    
    def create_summary_card(self, parent, title, value, color, row, col):
        """Özet kartı oluştur"""
        card_frame = ttk.LabelFrame(parent, text=title, padding=10)
        card_frame.grid(row=row, column=col, padx=10, pady=5, sticky='ew')
        
        value_label = tk.Label(card_frame, text=value, 
                              font=('Arial', 16, 'bold'), 
                              fg=color)
        value_label.pack()
        
        # Grid ağırlıklarını ayarla
        parent.columnconfigure(col, weight=1)
        
        # Kart referansını sakla
        setattr(self, f"card_{title.lower().replace(' ', '_')}", value_label)
    
    def generate_report(self):
        """Seçilen rapor türüne göre rapor oluştur"""
        report_type = self.report_type.get()
        date_range = self.date_range.get()
        
        # Tarih filtresi oluştur
        date_filter = self.get_date_filter(date_range)
        
        if report_type == "stock_summary":
            self.generate_stock_summary()
        elif report_type == "stock_movements":
            self.generate_stock_movements(date_filter)
        elif report_type == "low_stock":
            self.generate_low_stock_report()
        elif report_type == "value_analysis":
            self.generate_value_analysis()
        
        # Özet bilgileri güncelle
        self.update_summary_cards()
    
    def get_date_filter(self, date_range):
        """Tarih filtresi oluştur"""
        if date_range == "all":
            return ""
        elif date_range == "today":
            today = datetime.now().strftime('%Y-%m-%d')
            return f"AND DATE(movement_date) = '{today}'"
        elif date_range == "week":
            week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
            return f"AND DATE(movement_date) >= '{week_ago}'"
        elif date_range == "month":
            month_ago = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
            return f"AND DATE(movement_date) >= '{month_ago}'"
        return ""
    
    def generate_stock_summary(self):
        """Stok özet raporu"""
        # Tablo sütunlarını ayarla
        columns = ('Ürün Adı', 'Kategori', 'Barkod', 'Mevcut Stok', 'Min. Stok', 'Birim Fiyat', 'Toplam Değer', 'Durum')
        self.report_tree['columns'] = columns
        
        for col in columns:
            self.report_tree.heading(col, text=col)
            if col in ['Mevcut Stok', 'Min. Stok', 'Birim Fiyat', 'Toplam Değer']:
                self.report_tree.column(col, width=100)
            else:
                self.report_tree.column(col, width=120)
        
        # Verileri temizle
        for item in self.report_tree.get_children():
            self.report_tree.delete(item)
        
        # Verileri getir
        products = self.db.get_all_products()
        
        for product in products:
            total_value = product[5] * product[4]  # current_stock * unit_price
            status = "Kritik" if product[5] <= product[6] else "Normal"
            
            values = (product[1], product[3], product[2], product[5], 
                     product[6], f"{product[4]:.2f} ₺", f"{total_value:.2f} ₺", status)
            
            tags = ('critical',) if status == "Kritik" else ()
            self.report_tree.insert('', 'end', values=values, tags=tags)
        
        self.report_tree.tag_configure('critical', background='#ffcccc')
    
    def generate_stock_movements(self, date_filter):
        """Stok hareketleri raporu"""
        columns = ('Tarih', 'Ürün', 'İşlem', 'Miktar', 'Birim Fiyat', 'Toplam', 'Not')
        self.report_tree['columns'] = columns
        
        for col in columns:
            self.report_tree.heading(col, text=col)
            self.report_tree.column(col, width=120)
        
        # Verileri temizle
        for item in self.report_tree.get_children():
            self.report_tree.delete(item)
        
        # Verileri getir
        movements = self.db.get_stock_movements(date_filter)
        
        for movement in movements:
            date_str = movement[0][:16] if movement[0] else ''
            operation_text = 'Giriş' if movement[2] == 'IN' else 'Çıkış'
            
            values = (date_str, movement[1], operation_text, movement[3], 
                     f"{movement[4]:.2f} ₺", f"{movement[5]:.2f} ₺", movement[6] or '')
            
            tags = ('stock_in',) if movement[2] == 'IN' else ('stock_out',)
            self.report_tree.insert('', 'end', values=values, tags=tags)
        
        self.report_tree.tag_configure('stock_in', background='#e8f5e8')
        self.report_tree.tag_configure('stock_out', background='#ffe8e8')
    
    def generate_low_stock_report(self):
        """Düşük stok raporu"""
        columns = ('Ürün Adı', 'Kategori', 'Mevcut Stok', 'Min. Stok', 'Eksik Miktar', 'Tahmini Maliyet')
        self.report_tree['columns'] = columns
        
        for col in columns:
            self.report_tree.heading(col, text=col)
            self.report_tree.column(col, width=130)
        
        # Verileri temizle
        for item in self.report_tree.get_children():
            self.report_tree.delete(item)
        
        # Düşük stoklu ürünleri getir
        low_stock_products = self.db.get_low_stock_products()
        
        for product in low_stock_products:
            # Eksik miktarı hesapla
            missing_qty = product[3] - product[2]  # min_stock - current_stock
            
            # Tahmini maliyeti hesapla (ürün fiyatını al)
            product_details = self.db.execute_query(
                "SELECT unit_price FROM products WHERE id = ?", (product[0],)
            )
            unit_price = product_details[0][0] if product_details else 0
            estimated_cost = missing_qty * unit_price
            
            values = (product[1], product[4], product[2], product[3], 
                     missing_qty, f"{estimated_cost:.2f} ₺")
            
            self.report_tree.insert('', 'end', values=values, tags=('critical',))
        
        self.report_tree.tag_configure('critical', background='#ffcccc')
    
    def generate_value_analysis(self):
        """Değer analizi raporu"""
        columns = ('Ürün Adı', 'Kategori', 'Stok Miktarı', 'Birim Fiyat', 'Toplam Değer', 'Değer Oranı')
        self.report_tree['columns'] = columns
        
        for col in columns:
            self.report_tree.heading(col, text=col)
            self.report_tree.column(col, width=120)
        
        # Verileri temizle
        for item in self.report_tree.get_children():
            self.report_tree.delete(item)
        
        # Ürünleri getir ve toplam değeri hesapla
        products = self.db.get_all_products()
        total_inventory_value = sum(p[5] * p[4] for p in products)  # stock * price
        
        # Değere göre sırala
        product_values = []
        for product in products:
            total_value = product[5] * product[4]
            value_ratio = (total_value / total_inventory_value * 100) if total_inventory_value > 0 else 0
            product_values.append((product, total_value, value_ratio))
        
        # Değere göre azalan sırada sırala
        product_values.sort(key=lambda x: x[1], reverse=True)
        
        for product, total_value, value_ratio in product_values:
            values = (product[1], product[3], product[5], f"{product[4]:.2f} ₺", 
                     f"{total_value:.2f} ₺", f"{value_ratio:.1f}%")
            
            # Yüksek değerli ürünleri vurgula
            tags = ('high_value',) if value_ratio > 10 else ()
            self.report_tree.insert('', 'end', values=values, tags=tags)
        
        self.report_tree.tag_configure('high_value', background='#fff2cc')
    
    def update_summary_cards(self):
        """Özet kartlarını güncelle"""
        try:
            # Toplam ürün sayısı
            products = self.db.get_all_products()
            total_products = len(products)
            self.card_toplam_ürün.config(text=str(total_products))
            
            # Toplam stok değeri
            total_value = sum(p[5] * p[4] for p in products)
            self.card_toplam_değer.config(text=f"{total_value:.0f} ₺")
            
            # Düşük stoklu ürün sayısı
            low_stock = self.db.get_low_stock_products()
            self.card_düşük_stok.config(text=str(len(low_stock)))
            
            # Bu ay hareket sayısı
            month_filter = self.get_date_filter("month")
            monthly_movements = self.db.execute_query(f'''
                SELECT COUNT(*) FROM stock_movements 
                WHERE 1=1 {month_filter}
            ''')
            self.card_aylık_hareket.config(text=str(monthly_movements[0][0] if monthly_movements else 0))
            
            # En çok çıkış yapan ürün
            top_product = self.db.execute_query('''
                SELECT p.name, SUM(sm.quantity) as total_out
                FROM stock_movements sm
                JOIN products p ON sm.product_id = p.id
                WHERE sm.movement_type = 'OUT'
                GROUP BY p.id, p.name
                ORDER BY total_out DESC
                LIMIT 1
            ''')
            top_product_name = top_product[0][0] if top_product else "-"
            self.card_en_çok_satan.config(text=top_product_name[:15] + "..." if len(top_product_name) > 15 else top_product_name)
            
            # Kritik seviyedeki ürün sayısı (stok = 0)
            critical_products = self.db.execute_query('''
                SELECT COUNT(*) FROM products WHERE current_stock = 0
            ''')
            self.card_kritik_seviye.config(text=str(critical_products[0][0] if critical_products else 0))
            
            # Detaylı analiz metni
            self.update_detail_analysis(products, total_value, low_stock)
            
        except Exception as e:
            print(f"Özet kartları güncellenirken hata: {e}")
    
    def update_detail_analysis(self, products, total_value, low_stock):
        """Detaylı analiz metnini güncelle"""
        self.detail_text.delete(1.0, tk.END)
        
        analysis = f"""
📊 STOK ANALİZ RAPORU
{'='*50}

📈 GENEL DURUM:
• Toplam Ürün Çeşidi: {len(products)}
• Toplam Stok Değeri: {total_value:.2f} ₺
• Ortalama Ürün Değeri: {(total_value/len(products)):.2f} ₺

⚠️ UYARI DURUMU:
• Düşük Stoklu Ürün: {len(low_stock)}
• Kritik Seviye Oranı: {(len(low_stock)/len(products)*100):.1f}%

💰 DEĞER ANALİZİ:
"""
        
        # En değerli ürünleri bul
        valuable_products = sorted(products, key=lambda x: x[5] * x[4], reverse=True)[:5]
        analysis += "\n🏆 EN DEĞERLİ ÜRÜNLER:\n"
        for i, product in enumerate(valuable_products, 1):
            value = product[5] * product[4]
            analysis += f"{i}. {product[1]}: {value:.2f} ₺\n"
        
        # Kategori analizi
        categories = {}
        for product in products:
            cat = product[3] or "Kategorisiz"
            if cat not in categories:
                categories[cat] = {'count': 0, 'value': 0}
            categories[cat]['count'] += 1
            categories[cat]['value'] += product[5] * product[4]
        
        analysis += "\n📂 KATEGORİ ANALİZİ:\n"
        for cat, data in categories.items():
            analysis += f"• {cat}: {data['count']} ürün, {data['value']:.2f} ₺\n"
        
        # Öneriler
        analysis += f"\n💡 ÖNERİLER:\n"
        if len(low_stock) > 0:
            analysis += f"• {len(low_stock)} ürün için acil tedarik gerekli\n"
        
        avg_stock_ratio = sum(p[5]/max(p[6], 1) for p in products) / len(products)
        if avg_stock_ratio < 2:
            analysis += "• Genel stok seviyeleri düşük, tedarik planı gözden geçirilmeli\n"
        
        analysis += f"• Toplam stok değeri {total_value:.0f} ₺, nakit akışı planlaması yapılmalı\n"
        
        self.detail_text.insert(1.0, analysis)
    
    def export_to_excel(self):
        """Excel'e aktar"""
        try:
            # Dosya kaydetme dialogu
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Excel Dosyasını Kaydet",
                initialfile=f"Stok_Raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not file_path:
                return
            
            # Excel workbook oluştur
            wb = openpyxl.Workbook()
            
            # Mevcut sheet'i sil ve yeni sheet'ler oluştur
            wb.remove(wb.active)
            
            # Rapor türüne göre export
            report_type = self.report_type.get()
            
            if report_type == "stock_summary":
                self.export_stock_summary_to_excel(wb)
            elif report_type == "stock_movements":
                self.export_stock_movements_to_excel(wb)
            elif report_type == "low_stock":
                self.export_low_stock_to_excel(wb)
            elif report_type == "value_analysis":
                self.export_value_analysis_to_excel(wb)
            
            # Özet sayfası ekle
            self.add_summary_sheet_to_excel(wb)
            
            # Dosyayı kaydet
            wb.save(file_path)
            
            # Başarı mesajı ve dosyayı aç
            result = messagebox.askyesno("Başarılı", 
                                       f"Rapor başarıyla Excel'e aktarıldı!\n\n{file_path}\n\nDosyayı açmak ister misiniz?")
            
            if result:
                self.open_file(file_path)
                
        except Exception as e:
            messagebox.showerror("Hata", f"Excel'e aktarırken hata oluştu:\n{str(e)}")
    
    def export_stock_summary_to_excel(self, wb):
        """Stok özeti Excel'e aktar"""
        ws = wb.create_sheet("Stok Özeti")
        
        # Başlık
        ws['A1'] = "STOK ÖZETİ RAPORU"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        # Sütun başlıkları
        headers = ['Ürün Adı', 'Kategori', 'Barkod', 'Mevcut Stok', 'Min. Stok', 'Birim Fiyat', 'Toplam Değer', 'Durum']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Veriler
        products = self.db.get_all_products()
        for row, product in enumerate(products, 5):
            total_value = product[5] * product[4]
            status = "Kritik" if product[5] <= product[6] else "Normal"
            
            ws.cell(row=row, column=1, value=product[1])
            ws.cell(row=row, column=2, value=product[3])
            ws.cell(row=row, column=3, value=product[2])
            ws.cell(row=row, column=4, value=product[5])
            ws.cell(row=row, column=5, value=product[6])
            ws.cell(row=row, column=6, value=product[4])
            ws.cell(row=row, column=7, value=total_value)
            ws.cell(row=row, column=8, value=status)
            
            # Kritik ürünleri kırmızı yap
            if status == "Kritik":
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Sütun genişliklerini ayarla
        column_widths = [20, 15, 15, 12, 12, 12, 15, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    def export_stock_movements_to_excel(self, wb):
        """Stok hareketleri Excel'e aktar"""
        ws = wb.create_sheet("Stok Hareketleri")
        
        # Başlık
        ws['A1'] = "STOK HAREKETLERİ RAPORU"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        # Sütun başlıkları
        headers = ['Tarih', 'Ürün', 'İşlem', 'Miktar', 'Birim Fiyat', 'Toplam', 'Not']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Veriler
        date_filter = self.get_date_filter(self.date_range.get())
        query = f'''
            SELECT sm.movement_date, p.name, sm.movement_type, sm.quantity, 
                   sm.unit_price, sm.total_price, sm.note
            FROM stock_movements sm
            JOIN products p ON sm.product_id = p.id
            WHERE 1=1 {date_filter}
            ORDER BY sm.movement_date DESC
        '''
        movements = self.db.execute_query(query)
        
        for row, movement in enumerate(movements, 5):
            date_str = movement[0][:16] if movement[0] else ''
            operation_text = 'Giriş' if movement[2] == 'IN' else 'Çıkış'
            
            ws.cell(row=row, column=1, value=date_str)
            ws.cell(row=row, column=2, value=movement[1])
            ws.cell(row=row, column=3, value=operation_text)
            ws.cell(row=row, column=4, value=movement[3])
            ws.cell(row=row, column=5, value=movement[4])
            ws.cell(row=row, column=6, value=movement[5])
            ws.cell(row=row, column=7, value=movement[6] or '')
            
            # Giriş/çıkış renklendir
            color = "E8F5E8" if movement[2] == 'IN' else "FFE8E8"
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Sütun genişliklerini ayarla
        column_widths = [18, 20, 10, 10, 12, 12, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    def export_low_stock_to_excel(self, wb):
        """Düşük stok raporu Excel'e aktar"""
        ws = wb.create_sheet("Düşük Stok")
        
        # Başlık
        ws['A1'] = "DÜŞÜK STOK RAPORU"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        # Sütun başlıkları
        headers = ['Ürün Adı', 'Kategori', 'Mevcut Stok', 'Min. Stok', 'Eksik Miktar', 'Tahmini Maliyet']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Veriler
        low_stock_products = self.db.get_low_stock_products()
        
        for row, product in enumerate(low_stock_products, 5):
            missing_qty = product[3] - product[2]
            product_details = self.db.execute_query(
                "SELECT unit_price FROM products WHERE id = ?", (product[0],)
            )
            unit_price = product_details[0][0] if product_details else 0
            estimated_cost = missing_qty * unit_price
            
            ws.cell(row=row, column=1, value=product[1])
            ws.cell(row=row, column=2, value=product[4])
            ws.cell(row=row, column=3, value=product[2])
            ws.cell(row=row, column=4, value=product[3])
            ws.cell(row=row, column=5, value=missing_qty)
            ws.cell(row=row, column=6, value=estimated_cost)
            
            # Tüm satırı kırmızı yap
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Sütun genişliklerini ayarla
        column_widths = [20, 15, 12, 12, 12, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    def export_value_analysis_to_excel(self, wb):
        """Değer analizi Excel'e aktar"""
        ws = wb.create_sheet("Değer Analizi")
        
        # Başlık
        ws['A1'] = "DEĞER ANALİZİ RAPORU"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        # Sütun başlıkları
        headers = ['Ürün Adı', 'Kategori', 'Stok Miktarı', 'Birim Fiyat', 'Toplam Değer', 'Değer Oranı']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Veriler
        products = self.db.get_all_products()
        total_inventory_value = sum(p[5] * p[4] for p in products)
        
        product_values = []
        for product in products:
            total_value = product[5] * product[4]
            value_ratio = (total_value / total_inventory_value * 100) if total_inventory_value > 0 else 0
            product_values.append((product, total_value, value_ratio))
        
        product_values.sort(key=lambda x: x[1], reverse=True)
        
        for row, (product, total_value, value_ratio) in enumerate(product_values, 5):
            ws.cell(row=row, column=1, value=product[1])
            ws.cell(row=row, column=2, value=product[3])
            ws.cell(row=row, column=3, value=product[5])
            ws.cell(row=row, column=4, value=product[4])
            ws.cell(row=row, column=5, value=total_value)
            ws.cell(row=row, column=6, value=f"{value_ratio:.1f}%")
            
            # Yüksek değerli ürünleri vurgula
            if value_ratio > 10:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Sütun genişliklerini ayarla
        column_widths = [20, 15, 12, 12, 15, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    def add_summary_sheet_to_excel(self, wb):
        """Özet sayfası ekle"""
        ws = wb.create_sheet("Özet", 0)  # İlk sayfa olarak ekle
        
        # Başlık
        ws['A1'] = "STOK TAKİP SİSTEMİ - GENEL ÖZET"
        ws['A1'].font = Font(size=18, bold=True)
        ws['A2'] = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        # Özet bilgiler
        products = self.db.get_all_products()
        total_value = sum(p[5] * p[4] for p in products)
        low_stock = self.db.get_low_stock_products()
        
        summary_data = [
            ("Toplam Ürün Sayısı", len(products)),
            ("Toplam Stok Değeri", f"{total_value:.2f} ₺"),
            ("Düşük Stoklu Ürün", len(low_stock)),
            ("Kritik Seviye Oranı", f"{(len(low_stock)/len(products)*100):.1f}%" if products else "0%")
        ]
        
        for row, (label, value) in enumerate(summary_data, 4):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
        
        # En değerli ürünler
        ws['A9'] = "EN DEĞERLİ ÜRÜNLER (TOP 5)"
        ws['A9'].font = Font(size=14, bold=True)
        
        valuable_products = sorted(products, key=lambda x: x[5] * x[4], reverse=True)[:5]
        for row, product in enumerate(valuable_products, 10):
            value = product[5] * product[4]
            ws.cell(row=row, column=1, value=f"{row-9}. {product[1]}")
            ws.cell(row=row, column=2, value=f"{value:.2f} ₺")
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def print_report(self):
        """Raporu yazdır"""
        try:
            # Geçici HTML dosyası oluştur
            temp_file = os.path.join(os.path.expanduser("~"), "temp_report.html")
            
            # HTML içeriği oluştur
            html_content = self.generate_html_report()
            
            # HTML dosyasını yaz
            with open(temp_file, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            # Dosyayı varsayılan tarayıcıda aç (yazdırma için)
            self.open_file(temp_file)
            
            messagebox.showinfo("Bilgi", 
                              "Rapor tarayıcıda açıldı.\n\n"
                              "Yazdırmak için:\n"
                              "• Ctrl+P tuşlarına basın\n"
                              "• Veya Dosya > Yazdır menüsünü kullanın")
            
        except Exception as e:
            messagebox.showerror("Hata", f"Yazdırma hazırlığında hata oluştu:\n{str(e)}")
    
    def generate_html_report(self):
        """HTML rapor oluştur"""
        report_type = self.report_type.get()
        date_range = self.date_range.get()
        
        # Rapor başlığı
        report_titles = {
            "stock_summary": "Stok Özeti Raporu",
            "stock_movements": "Stok Hareketleri Raporu", 
            "low_stock": "Düşük Stok Raporu",
            "value_analysis": "Değer Analizi Raporu"
        }
        
        title = report_titles.get(report_type, "Stok Raporu")
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>{title}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #2c3e50; text-align: center; }}
                h2 {{ color: #34495e; border-bottom: 2px solid #3498db; padding-bottom: 5px; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #3498db; color: white; }}
                .critical {{ background-color: #ffcccc; }}
                .stock-in {{ background-color: #e8f5e8; }}
                .stock-out {{ background-color: #ffe8e8; }}
                .high-value {{ background-color: #fff2cc; }}
                .summary {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin: 20px 0; }}
                @media print {{
                    body {{ margin: 0; }}
                    .no-print {{ display: none; }}
                }}
            </style>
        </head>
        <body>
            <h1>STOK TAKİP SİSTEMİ</h1>
            <h2>{title}</h2>
            <p><strong>Rapor Tarihi:</strong> {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
            <p><strong>Tarih Aralığı:</strong> {self.get_date_range_text(date_range)}</p>
        """
        
        # Tablo verilerini ekle
        if report_type == "stock_summary":
            html += self.get_stock_summary_html()
        elif report_type == "stock_movements":
            html += self.get_stock_movements_html()
        elif report_type == "low_stock":
            html += self.get_low_stock_html()
        elif report_type == "value_analysis":
            html += self.get_value_analysis_html()
        
        # Özet bilgileri ekle
        html += self.get_summary_html()
        
        html += """
        </body>
        </html>
        """
        
        return html
    
    def get_date_range_text(self, date_range):
        """Tarih aralığı metnini döndür"""
        texts = {
            "all": "Tüm Kayıtlar",
            "today": "Bugün",
            "week": "Bu Hafta", 
            "month": "Bu Ay"
        }
        return texts.get(date_range, "Tüm Kayıtlar")
    
    def get_stock_summary_html(self):
        """Stok özeti HTML tablosu"""
        products = self.db.get_all_products()
        
        html = """
        <table>
            <tr>
                <th>Ürün Adı</th>
                <th>Kategori</th>
                <th>Barkod</th>
                <th>Mevcut Stok</th>
                <th>Min. Stok</th>
                <th>Birim Fiyat</th>
                <th>Toplam Değer</th>
                <th>Durum</th>
            </tr>
        """
        
        for product in products:
            total_value = product[5] * product[4]
            status = "Kritik" if product[5] <= product[6] else "Normal"
            css_class = "critical" if status == "Kritik" else ""
            
            html += f"""
            <tr class="{css_class}">
                <td>{product[1]}</td>
                <td>{product[3]}</td>
                <td>{product[2]}</td>
                <td>{product[5]}</td>
                <td>{product[6]}</td>
                <td>{product[4]:.2f} ₺</td>
                <td>{total_value:.2f} ₺</td>
                <td>{status}</td>
            </tr>
            """
        
        html += "</table>"
        return html
    
    def get_stock_movements_html(self):
        """Stok hareketleri HTML tablosu"""
        date_filter = self.get_date_filter(self.date_range.get())
        query = f'''
            SELECT sm.movement_date, p.name, sm.movement_type, sm.quantity, 
                   sm.unit_price, sm.total_price, sm.note
            FROM stock_movements sm
            JOIN products p ON sm.product_id = p.id
            WHERE 1=1 {date_filter}
            ORDER BY sm.movement_date DESC
            LIMIT 100
        '''
        movements = self.db.execute_query(query)
        
        html = """
        <table>
            <tr>
                <th>Tarih</th>
                <th>Ürün</th>
                <th>İşlem</th>
                <th>Miktar</th>
                <th>Birim Fiyat</th>
                <th>Toplam</th>
                <th>Not</th>
            </tr>
        """
        
        for movement in movements:
            date_str = movement[0][:16] if movement[0] else ''
            operation_text = 'Giriş' if movement[2] == 'IN' else 'Çıkış'
            css_class = "stock-in" if movement[2] == 'IN' else "stock-out"
            
            html += f"""
            <tr class="{css_class}">
                <td>{date_str}</td>
                <td>{movement[1]}</td>
                <td>{operation_text}</td>
                <td>{movement[3]}</td>
                <td>{movement[4]:.2f} ₺</td>
                <td>{movement[5]:.2f} ₺</td>
                <td>{movement[6] or ''}</td>
            </tr>
            """
        
        html += "</table>"
        return html
    
    def get_low_stock_html(self):
        """Düşük stok HTML tablosu"""
        low_stock_products = self.db.get_low_stock_products()
        
        html = """
        <table>
            <tr>
                <th>Ürün Adı</th>
                <th>Kategori</th>
                <th>Mevcut Stok</th>
                <th>Min. Stok</th>
                <th>Eksik Miktar</th>
                <th>Tahmini Maliyet</th>
            </tr>
        """
        
        for product in low_stock_products:
            missing_qty = product[3] - product[2]
            product_details = self.db.execute_query(
                "SELECT unit_price FROM products WHERE id = ?", (product[0],)
            )
            unit_price = product_details[0][0] if product_details else 0
            estimated_cost = missing_qty * unit_price
            
            html += f"""
            <tr class="critical">
                <td>{product[1]}</td>
                <td>{product[4]}</td>
                <td>{product[2]}</td>
                <td>{product[3]}</td>
                <td>{missing_qty}</td>
                <td>{estimated_cost:.2f} ₺</td>
            </tr>
            """
        
        html += "</table>"
        return html
    
    def get_value_analysis_html(self):
        """Değer analizi HTML tablosu"""
        products = self.db.get_all_products()
        total_inventory_value = sum(p[5] * p[4] for p in products)
        
        product_values = []
        for product in products:
            total_value = product[5] * product[4]
            value_ratio = (total_value / total_inventory_value * 100) if total_inventory_value > 0 else 0
            product_values.append((product, total_value, value_ratio))
        
        product_values.sort(key=lambda x: x[1], reverse=True)
        
        html = """
        <table>
            <tr>
                <th>Ürün Adı</th>
                <th>Kategori</th>
                <th>Stok Miktarı</th>
                <th>Birim Fiyat</th>
                <th>Toplam Değer</th>
                <th>Değer Oranı</th>
            </tr>
        """
        
        for product, total_value, value_ratio in product_values:
            css_class = "high-value" if value_ratio > 10 else ""
            
            html += f"""
            <tr class="{css_class}">
                <td>{product[1]}</td>
                <td>{product[3]}</td>
                <td>{product[5]}</td>
                <td>{product[4]:.2f} ₺</td>
                <td>{total_value:.2f} ₺</td>
                <td>{value_ratio:.1f}%</td>
            </tr>
            """
        
        html += "</table>"
        return html
    
    def get_summary_html(self):
        """Özet bilgileri HTML"""
        products = self.db.get_all_products()
        total_value = sum(p[5] * p[4] for p in products)
        low_stock = self.db.get_low_stock_products()
        
        html = f"""
        <div class="summary">
            <h2>Genel Özet</h2>
            <p><strong>Toplam Ürün Sayısı:</strong> {len(products)}</p>
            <p><strong>Toplam Stok Değeri:</strong> {total_value:.2f} ₺</p>
            <p><strong>Düşük Stoklu Ürün:</strong> {len(low_stock)}</p>
            <p><strong>Kritik Seviye Oranı:</strong> {(len(low_stock)/len(products)*100):.1f}%</p>
        </div>
        """
        
        return html
    
    def open_file(self, file_path):
        """Dosyayı sistem varsayılan programıyla aç"""
        try:
            if platform.system() == 'Windows':
                os.startfile(file_path)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', file_path])
            else:  # Linux
                subprocess.run(['xdg-open', file_path])
        except Exception as e:
            messagebox.showerror("Hata", f"Dosya açılırken hata oluştu:\n{str(e)}")